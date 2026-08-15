#!/usr/bin/env python3
"""Capability-checked Python backend for q-tool-spreadsheet."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from datetime import date, datetime
from pathlib import Path, PurePosixPath
from typing import Any
from xml.etree import ElementTree

MAX_ARCHIVE_BYTES = 64 * 1024 * 1024
MAX_EXPANDED_BYTES = 256 * 1024 * 1024
MAX_COMPRESSION_RATIO = 200
MAX_CELL_RECORDS = 1_000_000
MAX_MERGED_RANGES = 10_000
FORMULA_ERRORS = {"#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A"}
READ_ONLY_SUFFIXES = {".xlsm", ".xltm", ".xltx"}
WORKBOOK_SUFFIXES = {".xlsx", *READ_ONLY_SUFFIXES}
TABULAR_SUFFIXES = {".csv", ".tsv"}
SENSITIVE_PACKAGE_PREFIXES = {
    "xl/externallinks/": "external_links",
    "xl/connections": "data_connections",
    "xl/querytables/": "data_connections",
    "xl/embeddings/": "embedded_objects",
    "xl/activex/": "embedded_objects",
}
EXTERNAL_FORMULA = re.compile(
    r"(?i)(?:\b(?:WEBSERVICE|RTD|STOCKHISTORY|IMAGE)\s*\(|\[[^\]]+\][^!]*!|^\s*=\s*[A-Z0-9_.-]+\|[^!]+!)"
)


class ToolError(RuntimeError):
    pass


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bytes):
        return "<binary:%d>" % len(value)
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def emit(payload: dict[str, Any], as_json: bool) -> None:
    if as_json:
        print(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True))
    else:
        for key, value in payload.items():
            print("%s: %s" % (key, json.dumps(value, ensure_ascii=False)))


def openpyxl_status() -> tuple[bool, str | None]:
    try:
        import openpyxl  # type: ignore

        version = str(openpyxl.__version__)
        numbers = tuple(int(part) for part in re.findall(r"\d+", version)[:3])
        numbers = numbers + (0,) * (3 - len(numbers))
        return (3, 1, 5) <= numbers < (4, 0, 0), version
    except Exception:
        return False, None


def require_openpyxl():
    try:
        import openpyxl  # type: ignore
    except Exception as exc:
        raise ToolError(
            "Python backend requires openpyxl >=3.1.5,<4; the skill never installs dependencies"
        ) from exc
    available, version = openpyxl_status()
    if not available:
        raise ToolError("unsupported openpyxl version %s; require >=3.1.5,<4" % version)
    return openpyxl


def soffice_path() -> str | None:
    return shutil.which("soffice") or shutil.which("libreoffice")


def package_check(path: Path) -> dict[str, Any]:
    if not path.is_file():
        raise ToolError("input does not exist: %s" % path)
    if path.suffix.lower() in TABULAR_SUFFIXES:
        if path.stat().st_size > MAX_ARCHIVE_BYTES:
            raise ToolError("tabular input exceeds the bounded inspection size")
        return {
            "status": "passed",
            "kind": path.suffix.lower().lstrip("."),
            "entries": 1,
            "expanded_bytes": path.stat().st_size,
            "flags": [],
        }
    if path.suffix.lower() not in WORKBOOK_SUFFIXES:
        raise ToolError("unsupported spreadsheet type: %s" % path.suffix)
    if path.stat().st_size > MAX_ARCHIVE_BYTES:
        raise ToolError("workbook exceeds the bounded compressed size")

    flags: set[str] = set()
    names: set[str] = set()
    expanded = 0
    cell_records = 0
    merged_ranges = 0
    protected_worksheet_parts: list[str] = []
    workbook_protected = False
    try:
        with zipfile.ZipFile(path) as archive:
            for info in archive.infolist():
                name = info.filename.replace("\\", "/")
                pure = PurePosixPath(name)
                if pure.is_absolute() or ".." in pure.parts or name.startswith("/"):
                    raise ToolError("unsafe ZIP member path: %s" % info.filename)
                if info.flag_bits & 0x1:
                    raise ToolError("encrypted ZIP members are not supported")
                expanded += info.file_size
                if expanded > MAX_EXPANDED_BYTES:
                    raise ToolError("workbook exceeds the bounded expanded size")
                if info.compress_size and info.file_size / info.compress_size > MAX_COMPRESSION_RATIO:
                    raise ToolError("suspicious compression ratio in %s" % info.filename)
                if name in names:
                    raise ToolError("duplicate ZIP member name: %s" % name)
                names.add(name)
                lower = name.lower()
                if lower.endswith("vbaproject.bin"):
                    flags.add("macros")
                if lower.startswith("_xmlsignatures/"):
                    flags.add("signatures")
                for prefix, flag in SENSITIVE_PACKAGE_PREFIXES.items():
                    if lower.startswith(prefix):
                        flags.add(flag)
                if lower == "xl/workbook.xml" or (
                    lower.startswith("xl/worksheets/") and lower.endswith(".xml")
                ):
                    xml = archive.read(info)
                    if lower == "xl/workbook.xml" and re.search(
                        rb"<(?:[A-Za-z0-9_.-]+:)?workbookProtection\b", xml
                    ):
                        workbook_protected = True
                        flags.add("protection")
                    if lower.startswith("xl/worksheets/"):
                        cell_records += len(re.findall(rb"<(?:[A-Za-z0-9_.-]+:)?c(?:\s|>)", xml))
                        merged_ranges += len(re.findall(rb"<(?:[A-Za-z0-9_.-]+:)?mergeCell\b", xml))
                        if re.search(rb"<(?:[A-Za-z0-9_.-]+:)?sheetProtection\b", xml):
                            protected_worksheet_parts.append(name)
                            flags.add("protection")
                        if cell_records > MAX_CELL_RECORDS:
                            raise ToolError("workbook exceeds the one-million-cell inspection limit")
                        if merged_ranges > MAX_MERGED_RANGES:
                            raise ToolError("workbook exceeds the 10,000-merged-range inspection limit")
            required = {"[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"}
            missing = sorted(required - names)
            if missing:
                raise ToolError("missing required XLSX parts: %s" % ", ".join(missing))
            for name in sorted(required):
                try:
                    ElementTree.fromstring(archive.read(name))
                except Exception as exc:
                    raise ToolError("malformed required XML part %s: %s" % (name, exc)) from exc
    except zipfile.BadZipFile as exc:
        raise ToolError("invalid XLSX ZIP package") from exc

    return {
        "status": "passed",
        "kind": path.suffix.lower().lstrip("."),
        "entries": len(names),
        "expanded_bytes": expanded,
        "cell_records": cell_records,
        "merged_ranges": merged_ranges,
        "workbook_protected": workbook_protected,
        "protected_worksheet_parts": protected_worksheet_parts,
        "flags": sorted(flags),
    }


def workbook_inspect(path: Path, max_cells: int) -> dict[str, Any]:
    openpyxl = require_openpyxl()
    package = package_check(path)
    keep_vba = path.suffix.lower() in {".xlsm", ".xltm"}
    try:
        formulas = openpyxl.load_workbook(
            path, data_only=False, read_only=True, keep_vba=keep_vba, keep_links=True
        )
        cached = openpyxl.load_workbook(
            path, data_only=True, read_only=True, keep_vba=keep_vba, keep_links=True
        )
    except Exception as exc:
        raise ToolError("openpyxl could not inspect the workbook: %s" % exc) from exc

    formula_count = 0
    formula_errors: list[str] = []
    missing_cached: list[str] = []
    external_formulas: list[str] = []
    formula_records: list[str] = []
    sheets: list[dict[str, Any]] = []
    protected_sheets = list(package["protected_worksheet_parts"])
    preview: list[dict[str, Any]] = []
    populated_cells = 0
    try:
        for worksheet in formulas.worksheets:
            cached_sheet = cached[worksheet.title]
            sheet_formulas = 0
            for row, cached_row in zip(worksheet.iter_rows(), cached_sheet.iter_rows()):
                for cell, result_cell in zip(row, cached_row):
                    if cell.value is None:
                        continue
                    populated_cells += 1
                    if len(preview) < max_cells:
                        preview.append(
                            {
                                "sheet": worksheet.title,
                                "cell": cell.coordinate,
                                "value": json_value(cell.value),
                            }
                        )
                    if cell.data_type != "f":
                        continue
                    formula_count += 1
                    sheet_formulas += 1
                    formula_text = str(cell.value)
                    result = result_cell.value
                    location = "%s!%s" % (worksheet.title, cell.coordinate)
                    formula_records.append("%s=%s" % (location, formula_text))
                    if EXTERNAL_FORMULA.search(formula_text):
                        external_formulas.append(location)
                    if result_cell.data_type == "e" or result in FORMULA_ERRORS:
                        formula_errors.append(location)
                    elif result is None:
                        missing_cached.append(location)
            sheets.append(
                {
                    "name": worksheet.title,
                    "state": worksheet.sheet_state,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "formulas": sheet_formulas,
                }
            )
    finally:
        formulas.close()
        cached.close()

    return {
        "status": "passed",
        "runtime": "python",
        "file": str(path),
        "sha256": sha256(path),
        "package": package,
        "sheets": sheets,
        "workbook_protected": package["workbook_protected"],
        "protected_sheet_count": len(protected_sheets),
        "protected_sheets": protected_sheets,
        "formula_count": formula_count,
        "formula_sha256": hashlib.sha256("\n".join(formula_records).encode("utf-8")).hexdigest(),
        "formula_error_count": len(formula_errors),
        "formula_error_cells": formula_errors[:100],
        "missing_cached_values": len(missing_cached),
        "missing_cached_cells": missing_cached[:100],
        "external_formula_count": len(external_formulas),
        "external_formula_cells": external_formulas[:100],
        "preview": preview,
        "preview_truncated": populated_cells > len(preview),
    }


def require_distinct(input_path: Path, output_path: Path, overwrite: bool) -> None:
    try:
        same = input_path.resolve() == output_path.resolve()
    except FileNotFoundError:
        same = input_path.absolute() == output_path.absolute()
    if same:
        raise ToolError("input and output must be distinct")
    if output_path.exists() and not overwrite:
        raise ToolError("output already exists; replacement requires approval and --overwrite")
    output_path.parent.mkdir(parents=True, exist_ok=True)


def atomic_replace(source: Path, output: Path) -> None:
    os.replace(source, output)


def convert_tabular(input_path: Path, output_path: Path, sheet: str | None, view: str, overwrite: bool) -> dict[str, Any]:
    openpyxl = require_openpyxl()
    require_distinct(input_path, output_path, overwrite)
    package_check(input_path)
    source_suffix = input_path.suffix.lower()
    target_suffix = output_path.suffix.lower()
    if source_suffix in TABULAR_SUFFIXES and target_suffix == ".xlsx":
        delimiter = "\t" if source_suffix == ".tsv" else ","
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        title = (sheet or input_path.stem or "Sheet1")[:31]
        if not title or re.search(r"[\\/*?:\[\]]", title):
            raise ToolError("worksheet name contains an invalid XLSX character")
        worksheet.title = title
        with input_path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.reader(handle, delimiter=delimiter):
                worksheet.append(row)
                for cell in worksheet[worksheet.max_row]:
                    if isinstance(cell.value, str):
                        cell.data_type = "s"
        with tempfile.NamedTemporaryFile(
            prefix=".q-spreadsheet-", suffix=".xlsx", dir=output_path.parent, delete=False
        ) as temporary:
            temp_path = Path(temporary.name)
        try:
            workbook.save(temp_path)
            workbook.close()
            package_check(temp_path)
            atomic_replace(temp_path, output_path)
        finally:
            temp_path.unlink(missing_ok=True)
    elif source_suffix in WORKBOOK_SUFFIXES and target_suffix in TABULAR_SUFFIXES:
        if source_suffix in READ_ONLY_SUFFIXES:
            keep_vba = source_suffix in {".xlsm", ".xltm"}
        else:
            keep_vba = False
        workbook = openpyxl.load_workbook(
            input_path, data_only=(view == "values"), read_only=True, keep_vba=keep_vba, keep_links=True
        )
        try:
            worksheet = workbook[sheet] if sheet else workbook[workbook.sheetnames[0]]
            delimiter = "\t" if target_suffix == ".tsv" else ","
            with tempfile.NamedTemporaryFile(
                prefix=".q-spreadsheet-", suffix=target_suffix, dir=output_path.parent,
                delete=False, mode="w", encoding="utf-8", newline=""
            ) as temporary:
                temp_path = Path(temporary.name)
                writer = csv.writer(temporary, delimiter=delimiter, lineterminator="\n")
                for row in worksheet.iter_rows(values_only=True):
                    writer.writerow([json_value(value) if value is not None else "" for value in row])
            atomic_replace(temp_path, output_path)
        finally:
            workbook.close()
            if "temp_path" in locals():
                temp_path.unlink(missing_ok=True)
    else:
        raise ToolError("convert supports CSV/TSV to XLSX or XLSX-family input to CSV/TSV")
    return {
        "status": "completed",
        "runtime": "python",
        "operation": "convert",
        "input": str(input_path),
        "input_sha256": sha256(input_path),
        "output": str(output_path),
        "output_sha256": sha256(output_path),
    }


def safe_soffice_env(temp_root: Path) -> dict[str, str]:
    forwarded = (
        "PATH", "LANG", "LANGUAGE", "LC_ALL", "LC_CTYPE", "LC_NUMERIC", "LC_TIME",
        "SYSTEMROOT", "WINDIR", "COMSPEC", "PATHEXT", "SYSTEMDRIVE",
    )
    env = {name: os.environ[name] for name in forwarded if name in os.environ}
    env["HOME"] = str(temp_root)
    env["TMPDIR"] = str(temp_root)
    env["TMP"] = str(temp_root)
    env["TEMP"] = str(temp_root)
    env["SAL_USE_VCLPLUGIN"] = "svp"
    return env


def soffice_convert(input_path: Path, output_path: Path, format_name: str, timeout: int, overwrite: bool) -> str:
    require_distinct(input_path, output_path, overwrite)
    if timeout <= 0:
        raise ToolError("timeout must be a positive integer")
    if input_path.suffix.lower() != ".xlsx":
        raise ToolError("recalculation and rendering accept .xlsx only in this version")
    package = package_check(input_path)
    forbidden = {"macros", "external_links", "data_connections", "embedded_objects", "signatures", "protection"}
    active = sorted(forbidden.intersection(package.get("flags", [])))
    if active:
        raise ToolError("refusing local spreadsheet-engine execution for package flags: %s" % ", ".join(active))
    executable = soffice_path()
    if not executable:
        raise ToolError("LibreOffice is required for this operation; the skill never installs it")

    source_hash = sha256(input_path)
    temp_root = Path(tempfile.mkdtemp(prefix=".q-spreadsheet-", dir=output_path.parent))
    try:
        source_dir = temp_root / "source"
        converted_dir = temp_root / "converted"
        profile_dir = temp_root / "profile"
        source_dir.mkdir()
        converted_dir.mkdir()
        profile_dir.mkdir()
        staged_input = source_dir / "workbook.xlsx"
        shutil.copy2(input_path, staged_input)
        filter_spec = "xlsx:Calc MS Excel 2007 XML" if format_name == "xlsx" else "pdf:calc_pdf_Export"
        command = [
            executable,
            "--headless",
            "--nologo",
            "--nodefault",
            "--norestore",
            "--nolockcheck",
            "-env:UserInstallation=%s" % profile_dir.as_uri(),
            "--convert-to",
            filter_spec,
            "--outdir",
            str(converted_dir),
            str(staged_input),
        ]
        try:
            result = subprocess.run(
                command,
                text=True,
                capture_output=True,
                check=False,
                timeout=timeout,
                env=safe_soffice_env(temp_root),
            )
        except subprocess.TimeoutExpired as exc:
            raise ToolError("LibreOffice timed out; no output was committed") from exc
        if result.returncode != 0:
            detail = result.stderr.strip() or result.stdout.strip() or "exit %d" % result.returncode
            raise ToolError("LibreOffice conversion failed: %s" % detail)
        candidate = converted_dir / ("workbook.xlsx" if format_name == "xlsx" else "workbook.pdf")
        if not candidate.is_file() or candidate.stat().st_size == 0:
            raise ToolError("LibreOffice did not produce the expected non-empty output")
        if sha256(input_path) != source_hash:
            raise ToolError("source hash changed during conversion")
        atomic_replace(candidate, output_path)
        return result.stdout.strip()
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def recalculate(input_path: Path, output_path: Path, timeout: int, overwrite: bool) -> dict[str, Any]:
    require_distinct(input_path, output_path, overwrite)
    if output_path.suffix.lower() != ".xlsx":
        raise ToolError("recalculation output must use the .xlsx extension")
    before = workbook_inspect(input_path, max_cells=0)
    if before["external_formula_count"]:
        raise ToolError("refusing recalculation for formulas that may access external data")
    if before["workbook_protected"] or before["protected_sheet_count"]:
        raise ToolError("refusing recalculation for a protected workbook or worksheet")
    validation_root = Path(tempfile.mkdtemp(prefix=".q-spreadsheet-validation-", dir=output_path.parent))
    candidate = validation_root / "candidate.xlsx"
    try:
        log = soffice_convert(input_path, candidate, "xlsx", timeout, False)
        after = workbook_inspect(candidate, max_cells=0)
        blockers: list[str] = []
        warnings: list[str] = []
        if before["formula_count"] != after["formula_count"]:
            blockers.append("formula count changed during recalculation")
        if before["formula_sha256"] != after["formula_sha256"]:
            blockers.append("formula text changed during recalculation")
        if after["formula_error_count"]:
            blockers.append("recalculated workbook contains formula errors")
        if after["missing_cached_values"]:
            warnings.append("some formula cells have no cached value; blank-return formulas may be included")
        committed = not blockers
        if committed:
            atomic_replace(candidate, output_path)
        return {
            "status": "blocked" if blockers else ("completed_with_warnings" if warnings else "completed"),
            "runtime": "python",
            "operation": "recalculate",
            "input": str(input_path),
            "input_sha256": before["sha256"],
            "output": str(output_path) if committed else None,
            "output_sha256": sha256(output_path) if committed else None,
            "formula_count": after["formula_count"],
            "formula_sha256": after["formula_sha256"],
            "formula_error_count": after["formula_error_count"],
            "formula_error_cells": after["formula_error_cells"],
            "missing_cached_values": after["missing_cached_values"],
            "warnings": warnings,
            "blockers": blockers,
            "engine": "LibreOffice",
            "engine_log": log,
        }
    finally:
        shutil.rmtree(validation_root, ignore_errors=True)


def render(input_path: Path, output_dir: Path, timeout: int, overwrite: bool) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / (input_path.stem + ".pdf")
    inspection = workbook_inspect(input_path, max_cells=0)
    if inspection["external_formula_count"]:
        raise ToolError("refusing rendering for formulas that may access external data")
    if inspection["workbook_protected"] or inspection["protected_sheet_count"]:
        raise ToolError("refusing rendering for a protected workbook or worksheet")
    log = soffice_convert(input_path, output_path, "pdf", timeout, overwrite)
    return {
        "status": "completed",
        "runtime": "python",
        "operation": "render",
        "input": str(input_path),
        "input_sha256": sha256(input_path),
        "output": str(output_path),
        "output_sha256": sha256(output_path),
        "engine": "LibreOffice",
        "engine_log": log,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="q-tool-spreadsheet-python",
        description="Local Python backend for q-tool-spreadsheet; it never installs dependencies.",
        epilog="Mutating commands require distinct outputs and accept --overwrite only after separate approval.",
    )
    sub = parser.add_subparsers(dest="command", required=True)
    doctor = sub.add_parser("doctor", help="report backend capabilities")
    doctor.add_argument("--json", action="store_true")
    check = sub.add_parser("check", help="check bounded package safety and required parts")
    check.add_argument("input", type=Path)
    check.add_argument("--json", action="store_true")
    inspect = sub.add_parser("inspect", help="inspect workbook cells and formulas")
    inspect.add_argument("input", type=Path)
    inspect.add_argument("--max-cells", type=int, default=200)
    inspect.add_argument("--json", action="store_true")
    convert = sub.add_parser("convert", help="convert CSV/TSV and XLSX")
    convert.add_argument("input", type=Path)
    convert.add_argument("output", type=Path)
    convert.add_argument("--sheet")
    convert.add_argument("--view", choices=("values", "formulas"), default="values")
    convert.add_argument("--overwrite", action="store_true")
    convert.add_argument("--json", action="store_true")
    recalc = sub.add_parser("recalculate", help="recalculate into a distinct XLSX with LibreOffice")
    recalc.add_argument("input", type=Path)
    recalc.add_argument("output", type=Path)
    recalc.add_argument("--timeout", type=int, default=60)
    recalc.add_argument("--overwrite", action="store_true")
    recalc.add_argument("--json", action="store_true")
    render_parser = sub.add_parser("render", help="render XLSX to a validation PDF")
    render_parser.add_argument("input", type=Path)
    render_parser.add_argument("--output-dir", required=True, type=Path)
    render_parser.add_argument("--timeout", type=int, default=60)
    render_parser.add_argument("--overwrite", action="store_true")
    render_parser.add_argument("--json", action="store_true")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    as_json = bool(getattr(args, "json", False))
    try:
        if args.command == "doctor":
            available, version = openpyxl_status()
            executable = soffice_path()
            payload = {
                "status": "passed",
                "runtime": "python",
                "python": sys.version.split()[0],
                "openpyxl": available,
                "openpyxl_version": version,
                "soffice": bool(executable),
                "soffice_path": executable,
                "healthy": available,
                "safety_limits": {
                    "archive_bytes": MAX_ARCHIVE_BYTES,
                    "expanded_bytes": MAX_EXPANDED_BYTES,
                    "cell_records": MAX_CELL_RECORDS,
                    "merged_ranges": MAX_MERGED_RANGES,
                },
                "installs_dependencies": False,
            }
        elif args.command == "check":
            payload = package_check(args.input)
            payload["runtime"] = "python"
        elif args.command == "inspect":
            payload = workbook_inspect(args.input, max(0, args.max_cells))
        elif args.command == "convert":
            payload = convert_tabular(args.input, args.output, args.sheet, args.view, args.overwrite)
        elif args.command == "recalculate":
            payload = recalculate(args.input, args.output, args.timeout, args.overwrite)
        else:
            payload = render(args.input, args.output_dir, args.timeout, args.overwrite)
        emit(payload, as_json)
        return 4 if payload.get("status") == "blocked" else 0
    except ToolError as exc:
        payload = {"status": "blocked", "runtime": "python", "blockers": [str(exc)]}
        if as_json:
            emit(payload, True)
        else:
            print("Error: %s" % exc, file=sys.stderr)
        return 4
    except Exception as exc:
        payload = {"status": "blocked", "runtime": "python", "blockers": ["unexpected error: %s" % exc]}
        if as_json:
            emit(payload, True)
        else:
            print("Error: unexpected error: %s" % exc, file=sys.stderr)
        return 5


if __name__ == "__main__":
    raise SystemExit(main())
